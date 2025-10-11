import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def create_question_template(file_path: str):
    """创建题目导入模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "题目导入"
    
    # 设置表头
    headers = ["题目标题", "题目内容", "答案", "难度(1-5)", "学科ID", "知识点ID", "标签(用逗号分隔)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # 添加示例数据
    example = [
        "示例题目",
        "这是题目内容",
        "这是答案",
        "3",
        "1",
        "1",
        "基础,重点"
    ]
    for col, value in enumerate(example, 1):
        ws.cell(row=2, column=col, value=value)
    
    wb.save(file_path)
